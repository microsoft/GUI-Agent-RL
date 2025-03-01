import json
from typing import List
import openpyxl
from openpyxl.drawing.image import Image
import cv2
import pandas as pd
import matplotlib.pyplot as plt
import os
import math

from data_preprocess.utils import ActionType


def read_json(rpath: str):
    with open(rpath, 'r', encoding='utf-8') as file:
        data = json.load(file)
        return data


def write_json(anns: List, wpath: str):
    json.dump(anns, open(wpath, "w"))


def add_visilize2screenshot(image_rpath, ann, tag):
    if type(ann) == dict:
        if ann["action_type"] != "DUAL_POINT":
            return image_rpath

        touch_point, lift_point = ann["touch_point"], ann["lift_point"]
    else:
        if ann.action_type != ActionType.DualPoint:
            return image_rpath
        touch_point, lift_point = ann.touch_point, ann.lift_point

    click_point = [(touch_point[0] + lift_point[0]) / 2, (touch_point[1] + lift_point[1]) / 2]

    image = cv2.imread(image_rpath)
    height, width, _ = image.shape

    x = int(click_point[0] * width)
    y = int(click_point[1] * height)

    cv2.circle(image, (x, y), 20, (0, 0, 255), -1)

    image_wpath = image_rpath.split(".")[0] + f"_{tag}.png"
    cv2.imwrite(image_wpath, image) 

    return image_wpath.replace("\\", "/")


def write_to_excel(anns, wpath):
    wb = openpyxl.Workbook()
    ws = wb.active

    ws.cell(row=1, column=1, value="image")
    ws.cell(row=1, column=2, value="image(add point)")
    ws.cell(row=1, column=3, value="task")
    ws.cell(row=1, column=4, value="history action")
    ws.cell(row=1, column=5, value="current action")
    ws.cell(row=1, column=6, value="rating")
    ws.cell(row=1, column=7, value="explanation")

    for idx, ann in enumerate(anns, start=2):
        ws.cell(row=idx, column=3, value=ann["task"])
        ws.cell(row=idx, column=4, value="\n".join(ann["action_desc_list"]))
        ws.cell(row=idx, column=5, value=ann["action_desc_list"][ann["step_id"]])
        ws.cell(row=idx, column=6, value=ann["rating"])
        ws.cell(row=idx, column=7, value=ann["explanation"])

        img = Image(ann["image_list"][ann["step_id"]].replace("\\", "/"))
        img.width, img.height = (240, 480)
        ws.row_dimensions[idx].height = 400
        ws.add_image(img, f'A{idx}')
        img = Image(ann["add_point_image_list"][ann["step_id"]])
        img.width, img.height = (240, 480)
        ws.row_dimensions[idx].height = 400
        ws.add_image(img, f'B{idx}')


    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20
    wb.save(wpath)


def parse_response(response):
    try:
        response = response.replace("```", "").replace("json", "")
        return json.loads(response)
    except:
        return -1
    

def write_jsonl(anns, wpath):
    with open(wpath, 'w', encoding='utf - 8') as f:
        for item in anns:
            json_line = json.dumps(item)
            f.write(json_line + '\n')


def read_jsonl(rpath):
    data = []
    with open(rpath, 'r', encoding='utf - 8') as f:
        for idx, line in enumerate(f):
            line = line.strip()
            try:
                json_obj = json.loads(line)
                data.append(json_obj)
            except:
                print(f"Error decoding JSON on line: {idx}")
    return data


def read_xlsx(rpath):
    data = pd.read_excel(rpath)
    return data.to_dict(orient="records")


def dict_mean(dict_list):
    mean_dict = {}
    if len(dict_list) > 0:
        for key in dict_list[0].keys():
            if "min" in key:
                mean_dict[key] = min(d[key] for d in dict_list)
            elif "max" in key:
                mean_dict[key] = max(d[key] for d in dict_list)
            else:
                mean_dict[key] = sum(d[key] for d in dict_list) / len(dict_list)
    return mean_dict


def smooth(scalars: List[float]) -> List[float]:
    if len(scalars) == 0:
        return []

    last = scalars[0]
    smoothed = []
    weight = 1.9 * (1 / (1 + math.exp(-0.05 * len(scalars))) - 0.5)  # a sigmoid function
    for next_val in scalars:
        smoothed_val = last * weight + (1 - weight) * next_val
        smoothed.append(smoothed_val)
        last = smoothed_val
    return smoothed


def plot_loss(log_dir: str, keys: List[str] = ["loss"]) -> None:
    plt.switch_backend("agg")
    data = read_jsonl(os.path.join(log_dir, "train_log.jsonl"))

    for key in keys:
        steps, metrics = [], []
        for i in range(len(data)):
            if key in data[i]:
                steps.append(data[i]["step"])
                metrics.append(data[i][key])

        plt.figure()
        plt.plot(steps, metrics, color="#1f77b4", alpha=0.4, label="original")
        plt.plot(steps, smooth(metrics), color="#1f77b4", label="smoothed")
        plt.title(f"{key} of {log_dir}")
        plt.xlabel("step")
        plt.ylabel(key)
        plt.legend()
        figure_path = os.path.join(log_dir, "training_{}.png".format(key))
        plt.savefig(figure_path, format="png", dpi=100)
        print("Figure saved at:", figure_path)
