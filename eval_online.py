import argparse
import yaml
import json
import os
import cv2
import openpyxl
from openpyxl.drawing.image import Image
import cv2

import utils
from dataset import create_dataset
from models import create_agent
from data_preprocess.utils import ActionType, to_autoui
from eval_tools.androidenv import AndroidEnv


def add_visilize2screenshot(image_rpath, action):
    if action.action_type != ActionType.DualPoint:
        return image_rpath

    image = cv2.imread(image_rpath)
    height, width, _ = image.shape

    x = int(action.touch_point[0] * width)
    y = int(action.touch_point[1] * height)

    cv2.circle(image, (x, y), 50, (0, 0, 255), -1)

    image_wpath = image_rpath.replace(".png", "") + f"_point.png"
    cv2.imwrite(image_wpath, image) 

    return image_wpath


def evaluation(config, agent, dataset, env, ann_wpath):
    with open(ann_wpath, "a") as fout:
        for task_id, task, query_format in dataset:
            done, history = False, []

            step_num = 0
            current_screenshot_path = env.get_obs(task_id, step_num)
            while not done:
                step_num += 1
                text = query_format.format("\n".join(history), task)
                
                raw_action = agent.get_action(text=text, image_path=current_screenshot_path)
                action = env.translate_action(raw_action)
                point_image_path = add_visilize2screenshot(current_screenshot_path, action)

                next_screenshot_path, done, action, explanation = env.step(task_id, step_num, task, raw_action)
                
                action_desc = to_autoui(action, all_dict=True)
                history.append(action_desc)
                
                print("============")
                print(f"{task_id}: {task}")
                print(f"current_image: {current_screenshot_path}")
                print(f"action: {action}")
                print(f"point action on image: {point_image_path}")
                print(f"next screen shot: {next_screenshot_path}")
                print(f"if done: {done}")
                
                result ={
                    "task_id": task_id,
                    "step_id": step_num,
                    "task": task,
                    "action": history[-1],
                    "current_image_path": current_screenshot_path.replace("\\", "/"),
                    "point_image_path": point_image_path.replace("\\", "/"),
                    "next_image_path": next_screenshot_path.replace("\\", "/"),
                    "if_done": done,
                    "prompt": text,
                    "gpt-4o": explanation
                }

                fout.writelines(json.dumps(result) + "\n")

                current_screenshot_path = next_screenshot_path

                if step_num > 10 or done or action.action_type == ActionType.TaskComplete:
                    env.driver.press_keycode(3)
                    break
                
    fout.close()


def main(config):
    print("config:", json.dumps(config))
    ann_wpath = f"./data/{config['output_name']}.jsonl"
    finish_task, success_num, step_len = {}, 0, 0
    if os.path.exists(ann_wpath):
        for ann in utils.read_jsonl(ann_wpath):
            if ann["task"] not in finish_task:
                finish_task[ann["task"]] = {"success": False, "steps": []}
            if ann["if_done"]:
                finish_task[ann["task"]]["success"] = True
            finish_task[ann["task"]]["steps"].append(ann)

        for _, info in finish_task.items():
            if info["success"]: success_num += 1
            step_len += len(info["steps"])

        utils.write_json({"success_num": success_num, "step_num": step_len, "info": finish_task}, ann_wpath.replace("jsonl", "json"))
        if len(finish_task.keys()) > 0:
            print(f"### finish task num: {len(finish_task.keys())}\tsuccess: {success_num}\tstep_len: {step_len/len(finish_task.keys())}")

    print("output_path: ", ann_wpath)

    print("Creating datasets")
    dataset = create_dataset(config, finish_task)

    print("build android env")

    env = AndroidEnv(config)

    print("Creating agent")
    agent = create_agent(config)

    print("### Start evaluating")

    evaluation(config, agent, dataset, env, ann_wpath)


def write_to_excel(anns, wpath):
    wb = openpyxl.Workbook()
    ws = wb.active

    ws.cell(row=1, column=1, value="current image")
    ws.cell(row=1, column=2, value="current image(add point)")
    ws.cell(row=1, column=3, value="next image")
    ws.cell(row=1, column=4, value="task")
    ws.cell(row=1, column=5, value="action")
    ws.cell(row=1, column=6, value="prompt")
    ws.cell(row=1, column=7, value="if_done")
    ws.cell(row=1, column=8, value="explanation")

    for idx, ann in enumerate(anns, start=2):
        ws.cell(row=idx, column=4, value=ann["task"])
        ws.cell(row=idx, column=5, value=ann["action"])
        ws.cell(row=idx, column=6, value=ann["prompt"])
        ws.cell(row=idx, column=7, value=ann["if_done"])
        ws.cell(row=idx, column=8, value=ann["gpt-4o"])
        
        img = Image(ann["current_image_path"].replace("\\", "/"))
        img.width, img.height = (240, 480)
        ws.row_dimensions[idx].height = 400
        ws.add_image(img, f'A{idx}')
        img = Image(ann["point_image_path"])
        img.width, img.height = (240, 480)
        ws.row_dimensions[idx].height = 400
        ws.add_image(img, f'B{idx}')
        img = Image(ann["next_image_path"])
        img.width, img.height = (240, 480)
        ws.row_dimensions[idx].height = 400
        ws.add_image(img, f'C{idx}')


    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20
    wb.save(wpath)


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument('--task', type=str, required=True)

    args = parser.parse_args()

    config = f"configs/android_eval/online_eval_{args.task}.yaml"
    with open(config, 'r') as file:
        config = yaml.safe_load(file)

    main(config)